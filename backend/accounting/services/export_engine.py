"""
Export Engine - Excel and PDF export for reports.
Provides unified export interface for all report types.
"""
import io
import json
import base64
from datetime import date
from decimal import Decimal
from typing import Optional, Dict, Any
from abc import ABC, abstractmethod


class BaseExporter(ABC):
    """Base class for export implementations."""
    
    @abstractmethod
    def export(self, data: dict, report_type: str, company_name: str = '') -> bytes:
        """Export report data to format."""
        pass
    
    @staticmethod
    def _fmt(amount) -> str:
        """Format decimal amount."""
        if amount is None:
            return '0.00'
        if isinstance(amount, Decimal):
            return f'{amount:,.2f}'
        try:
            return f'{float(amount):,.2f}'
        except (ValueError, TypeError):
            return '0.00'


class ExcelExporter(BaseExporter):
    """Export reports to Excel format using openpyxl."""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.current_row = 1
    
    def export(self, data: dict, report_type: str, company_name: str = '') -> bytes:
        """Export to Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter as _get_column_letter
            self.get_column_letter = _get_column_letter

            self.workbook = Workbook()
            self.workbook.remove(self.workbook.active)
            
            # Add styles
            self.Alignment = Alignment
            self.header_font = Font(bold=True, size=12)
            self.title_font = Font(bold=True, size=14)
            self.bold_font = Font(bold=True)
            self.header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            self.alternate_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            
            if report_type == 'trial_balance':
                self._export_trial_balance(data, company_name)
            elif report_type == 'profit_loss':
                self._export_profit_loss(data, company_name)
            elif report_type == 'balance_sheet':
                self._export_balance_sheet(data, company_name)
            elif report_type == 'ledger':
                self._export_ledger(data, company_name)
            elif report_type == 'ar_aging':
                self._export_ar_aging(data, company_name)
            elif report_type == 'ap_aging':
                self._export_ap_aging(data, company_name)
            elif report_type == 'cash_flow':
                self._export_cash_flow(data, company_name)
            else:
                self._export_generic(data, company_name)
            
            output = io.BytesIO()
            self.workbook.save(output)
            return output.getvalue()
            
        except ImportError:
            # Fallback to CSV
            return self._fallback_csv(data, report_type, company_name)
    
    def _set_alignment(self, cell, horizontal='center', vertical=None):
        """Set cell alignment."""
        cell.alignment = self.Alignment(horizontal=horizontal)

    def _add_title(self, title: str, subtitle: str = ''):
        """Add title row."""
        ws = self.worksheet
        ws.merge_cells(f'A{self.current_row}:H{self.current_row}')
        cell = ws[f'A{self.current_row}']
        cell.value = title
        cell.font = self.title_font
        self._set_alignment(cell)
        self.current_row += 1
        
        if subtitle:
            ws.merge_cells(f'A{self.current_row}:H{self.current_row}')
            cell = ws[f'A{self.current_row}']
            cell.value = subtitle
            self._set_alignment(cell)
            self.current_row += 1
        self.current_row += 1
    
    def _add_header(self, headers: list):
        """Add header row with styling."""
        ws = self.worksheet
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=self.current_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            self._set_alignment(cell)
        self.current_row += 1
    
    def _add_row(self, values: list, bold: bool = False):
        """Add data row."""
        ws = self.worksheet
        fill = self.alternate_fill if self.current_row % 2 else None
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=self.current_row, column=col)
            if isinstance(value, (int, float, Decimal)):
                cell.value = float(value)
                cell.alignment = self.Alignment(horizontal='right')
            else:
                cell.value = str(value) if value else ''
            if bold:
                cell.font = self.bold_font
            if fill:
                cell.fill = fill
        self.current_row += 1
    
    def _export_trial_balance(self, data: dict, company_name: str):
        """Export Trial Balance."""
        ws = self.workbook.create_sheet('Trial Balance')
        self.worksheet = ws
        
        self._add_title(
            f'{company_name or "Pharmacy ERP"} - Trial Balance',
            f'As of: {data.get("as_of_date", date.today())}'
        )
        
        self._add_header(['Account Code', 'Account Name', 'Type', 'Category', 'Debit', 'Credit', 'Net Balance', 'Balance Type'])
        
        for row in data.get('accounts', []):
            self._add_row([
                row.get('account_code', ''),
                row.get('account_name', ''),
                row.get('account_type', ''),
                row.get('account_category', ''),
                self._fmt(row.get('total_debit', 0)),
                self._fmt(row.get('total_credit', 0)),
                self._fmt(row.get('net_balance', 0)),
                row.get('balance_type', '')
            ])
        
        self.current_row += 1
        self._add_row([
            '', '', '', 'TOTALS',
            self._fmt(data.get('total_debit', 0)),
            self._fmt(data.get('total_credit', 0)),
            '',
            'Balanced' if data.get('is_balanced') else 'NOT BALANCED'
        ], bold=True)
        
        self._auto_size_columns(8)
    
    def _export_profit_loss(self, data: dict, company_name: str):
        """Export Profit & Loss."""
        ws = self.workbook.create_sheet('Profit & Loss')
        self.worksheet = ws
        
        period = f"{data.get('start_date', '')} to {data.get('end_date', '')}"
        self._add_title(
            f'{company_name or "Pharmacy ERP"} - Profit & Loss Statement',
            f'Period: {period}'
        )
        
        self._add_header(['Description', 'Amount'])
        
        self._add_row(['REVENUE'], bold=True)
        for section in data.get('revenue', []):
            if isinstance(section, dict):
                self._add_row([section.get('category', ''), self._fmt(section.get('total', 0))], bold=True)
                for acc in section.get('accounts', []):
                    self._add_row([f"  {acc.get('account_code')} - {acc.get('account_name')}", self._fmt(acc.get('amount', 0))])
        self._add_row(['Total Revenue', self._fmt(data.get('total_revenue', 0))], bold=True)
        
        self.current_row += 1
        self._add_row(['COGS'], bold=True)
        for section in data.get('cogs', []):
            if isinstance(section, dict):
                for acc in section.get('accounts', []):
                    self._add_row([f"  {acc.get('account_code')} - {acc.get('account_name')}", self._fmt(acc.get('amount', 0))])
        self._add_row(['Total COGS', self._fmt(data.get('total_cogs', 0))], bold=True)
        
        self.current_row += 1
        self._add_row(['Gross Profit', self._fmt(data.get('gross_profit', 0))], bold=True)
        
        self.current_row += 1
        self._add_row(['EXPENSES'], bold=True)
        for section in data.get('expenses', []):
            if isinstance(section, dict):
                self._add_row([section.get('category', ''), self._fmt(section.get('total', 0))], bold=True)
                for acc in section.get('accounts', []):
                    self._add_row([f"  {acc.get('account_code')} - {acc.get('account_name')}", self._fmt(acc.get('amount', 0))])
        self._add_row(['Total Expenses', self._fmt(data.get('total_expenses', 0))], bold=True)
        
        self.current_row += 1
        self._add_row(['NET INCOME', self._fmt(data.get('net_income', 0))], bold=True)
        
        self._auto_size_columns(2)
    
    def _export_balance_sheet(self, data: dict, company_name: str):
        """Export Balance Sheet."""
        ws = self.workbook.create_sheet('Balance Sheet')
        self.worksheet = ws
        
        self._add_title(
            f'{company_name or "Pharmacy ERP"} - Balance Sheet',
            f'As of: {data.get("as_of_date", date.today())}'
        )
        
        self._add_header(['Section', 'Category', 'Account Code', 'Account Name', 'Amount'])
        
        for section in data.get('assets', {}).get('sections', []):
            self._add_row(['ASSETS', section.get('category', ''), '', '', self._fmt(section.get('total', 0))], bold=True)
            for acc in section.get('accounts', []):
                self._add_row(['', '', acc.get('account_code'), acc.get('account_name'), self._fmt(acc.get('amount', 0))])
        self._add_row(['', '', '', 'Total Assets', self._fmt(data.get('assets', {}).get('total', 0))], bold=True)
        
        self.current_row += 1
        for section in data.get('liabilities', {}).get('sections', []):
            self._add_row(['LIABILITIES', section.get('category', ''), '', '', self._fmt(section.get('total', 0))], bold=True)
            for acc in section.get('accounts', []):
                self._add_row(['', '', acc.get('account_code'), acc.get('account_name'), self._fmt(acc.get('amount', 0))])
        self._add_row(['', '', '', 'Total Liabilities', self._fmt(data.get('liabilities', {}).get('total', 0))], bold=True)
        
        self.current_row += 1
        for section in data.get('equity', {}).get('sections', []):
            self._add_row(['EQUITY', section.get('category', ''), '', '', self._fmt(section.get('total', 0))], bold=True)
            for acc in section.get('accounts', []):
                self._add_row(['', '', acc.get('account_code'), acc.get('account_name'), self._fmt(acc.get('amount', 0))])
        self._add_row(['', '', '', 'Total Equity', self._fmt(data.get('equity', {}).get('total', 0))], bold=True)
        
        self._auto_size_columns(5)
    
    def _export_ledger(self, data: dict, company_name: str):
        """Export Account Ledger."""
        ws = self.workbook.create_sheet('Ledger')
        self.worksheet = ws
        
        self._add_title(
            f'{company_name or "Pharmacy ERP"} - Account Ledger',
            f'Account: {data.get("account_code", "")} - {data.get("account_name", "")}'
        )
        
        self._add_header(['Entry Number', 'Date', 'Type', 'Description', 'Reference', 'Debit', 'Credit', 'Running Balance'])
        
        for entry in data.get('entries', []):
            self._add_row([
                entry.get('entry_number', ''),
                entry.get('entry_date', ''),
                entry.get('entry_type', ''),
                entry.get('description', ''),
                entry.get('reference', ''),
                self._fmt(entry.get('debit', 0)),
                self._fmt(entry.get('credit', 0)),
                self._fmt(entry.get('running_balance', 0))
            ])
        
        self.current_row += 1
        self._add_row([
            '', '', '', '', 'Closing Balance',
            '', '', self._fmt(data.get('closing_balance', 0))
        ], bold=True)
        
        self._auto_size_columns(8)
    
    def _export_ar_aging(self, data: dict, company_name: str):
        """Export AR Aging."""
        ws = self.workbook.create_sheet('AR Aging')
        self.worksheet = ws
        
        self._add_title(
            f'{company_name or "Pharmacy ERP"} - Accounts Receivable Aging',
            f'As of: {data.get("as_of_date", date.today())}'
        )
        
        self._add_header(['Customer', 'Current', '1-30 Days', '31-60 Days', '61-90 Days', '90+ Days', 'Total'])
        
        for row in data.get('aging_rows', []):
            self._add_row([
                row.get('party_name', ''),
                self._fmt(row.get('current', 0)),
                self._fmt(row.get('age_1_30', 0)),
                self._fmt(row.get('age_31_60', 0)),
                self._fmt(row.get('age_61_90', 0)),
                self._fmt(row.get('age_90_plus', 0)),
                self._fmt(row.get('total', 0))
            ])
        
        totals = data.get('totals', {})
        self.current_row += 1
        self._add_row([
            'TOTAL',
            self._fmt(totals.get('current', 0)),
            self._fmt(totals.get('age_1_30', 0)),
            self._fmt(totals.get('age_31_60', 0)),
            self._fmt(totals.get('age_61_90', 0)),
            self._fmt(totals.get('age_90_plus', 0)),
            self._fmt(totals.get('total', 0))
        ], bold=True)
        
        self._auto_size_columns(7)
    
    def _export_ap_aging(self, data: dict, company_name: str):
        """Export AP Aging."""
        self._export_ar_aging(data, company_name)  # Same structure
    
    def _export_cash_flow(self, data: dict, company_name: str):
        """Export Cash Flow."""
        ws = self.workbook.create_sheet('Cash Flow')
        self.worksheet = ws
        
        self._add_title(
            f'{company_name or "Pharmacy ERP"} - Cash Flow Statement',
            f'Period: {data.get("start_date", "")} to {data.get("end_date", "")}'
        )
        
        self._add_header(['Description', 'Amount'])
        
        op = data.get('operating_activities', {})
        self._add_row(['OPERATING ACTIVITIES'], bold=True)
        self._add_row(['Net Income', self._fmt(op.get('net_income', 0))])
        for item in op.get('working_capital_changes', []):
            self._add_row([f"  {item.get('description', '')}", self._fmt(item.get('change', 0))])
        self._add_row(['Net Cash from Operations', self._fmt(op.get('net_cash_from_operations', 0))], bold=True)
        
        self.current_row += 1
        inv = data.get('investing_activities', {})
        self._add_row(['INVESTING ACTIVITIES'], bold=True)
        for item in inv.get('items', []):
            self._add_row([item.get('description', ''), self._fmt(item.get('change', 0))])
        self._add_row(['Net Cash from Investing', self._fmt(inv.get('net_cash_from_investing', 0))], bold=True)
        
        self.current_row += 1
        fin = data.get('financing_activities', {})
        self._add_row(['FINANCING ACTIVITIES'], bold=True)
        for item in fin.get('items', []):
            self._add_row([item.get('description', ''), self._fmt(item.get('change', 0))])
        self._add_row(['Net Cash from Financing', self._fmt(fin.get('net_cash_from_financing', 0))], bold=True)
        
        self.current_row += 1
        self._add_row(['Net Change in Cash', self._fmt(data.get('net_change_in_cash', 0))], bold=True)
        self._add_row(['Opening Cash Balance', self._fmt(data.get('opening_cash_balance', 0))])
        self._add_row(['Closing Cash Balance', self._fmt(data.get('closing_cash_balance', 0))], bold=True)
        
        self._auto_size_columns(2)
    
    def _export_generic(self, data: dict, company_name: str):
        """Export generic data."""
        ws = self.workbook.create_sheet('Report')
        self.worksheet = ws
        
        self._add_title(f'{company_name or "Pharmacy ERP"} - Report', f'Generated: {date.today()}')
        
        for key, value in data.items():
            if not isinstance(value, (list, dict)):
                self._add_row([key, str(value)])
        
        self._auto_size_columns(2)
    
    def _auto_size_columns(self, num_cols: int):
        """Auto-size columns."""
        try:
            col_letter = self.get_column_letter
        except AttributeError:
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter
        for col in range(1, num_cols + 1):
            self.worksheet.column_dimensions[col_letter(col)].width = 15
    
    def _fallback_csv(self, data: dict, report_type: str, company_name: str) -> bytes:
        """Fallback to CSV if openpyxl not available."""
        from accounting.services.report_exporter import ReportExporter
        csv_content = ReportExporter.to_csv(data, report_type)
        return csv_content.encode('utf-8')


class PDFExporter(BaseExporter):
    """Export reports to PDF format using ReportLab."""
    
    def export(self, data: dict, report_type: str, company_name: str = '') -> bytes:
        """Export to PDF."""
        try:
            from reportlab.lib.pagesizes import A4, letter
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.pdfgen import canvas
            from io import BytesIO
            
            # Store reportlab references for use in _build_* methods
            self._rl_colors = colors
            self._rl_inch = inch
            self._rl_Paragraph = Paragraph
            self._rl_Spacer = Spacer
            self._rl_Table = Table
            self._rl_TableStyle = TableStyle

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
            story = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, spaceAfter=12)
            heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=12, spaceBefore=12, spaceAfter=6)
            normal_style = styles['Normal']
            
            # Title
            story.append(Paragraph(f'{company_name or "Pharmacy ERP"}', title_style))
            story.append(Paragraph(data.get('report_type', 'Report').title(), title_style))
            
            if 'as_of_date' in data:
                story.append(Paragraph(f'As of: {data["as_of_date"]}', normal_style))
            if 'start_date' in data and 'end_date' in data:
                story.append(Paragraph(f'Period: {data["start_date"]} to {data["end_date"]}', normal_style))
            
            story.append(Spacer(1, 0.25*inch))
            
            # Content based on report type
            if report_type == 'trial_balance':
                self._build_trial_balance_table(story, data, normal_style)
            elif report_type == 'profit_loss':
                self._build_profit_loss_table(story, data, normal_style)
            elif report_type == 'balance_sheet':
                self._build_balance_sheet_table(story, data, normal_style)
            elif report_type == 'ar_aging' or report_type == 'ap_aging':
                self._build_aging_table(story, data, normal_style)
            
            # Footer
            story.append(Spacer(1, 0.5*inch))
            story.append(Paragraph(f'Generated: {date.today()}', normal_style))
            
            # QR Code for verification
            from accounting.services.report_exporter import ReportExporter
            qr_b64 = ReportExporter.generate_qr_code_base64(data, report_type)
            if qr_b64:
                from reportlab.platypus import Image
                from io import BytesIO
                qr_data = BytesIO(base64.b64decode(qr_b64))
                qr_img = Image(qr_data, width=60, height=60)
                story.append(Spacer(1, 0.15*inch))
                story.append(qr_img)
                story.append(Paragraph('Scan to verify report', normal_style))
            
            doc.build(story)
            return buffer.getvalue()
            
        except ImportError:
            # Fallback to text
            from accounting.services.report_exporter import ReportExporter
            text_content = ReportExporter.to_text(data, report_type, company_name)
            return text_content.encode('utf-8')
    
    def _build_trial_balance_table(self, story, data, style):
        """Build trial balance table."""
        Table = self._rl_Table
        TableStyle = self._rl_TableStyle
        colors = self._rl_colors

        table_data = [['Account Code', 'Account Name', 'Debit', 'Credit']]
        
        for row in data.get('accounts', []):
            table_data.append([
                row.get('account_code', ''),
                row.get('account_name', ''),
                self._fmt(row.get('total_debit', 0)),
                self._fmt(row.get('total_credit', 0))
            ])
        
        table_data.append(['', 'TOTALS', self._fmt(data.get('total_debit', 0)), self._fmt(data.get('total_credit', 0))])
        
        t = Table(table_data, colWidths=[80, 200, 80, 80])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(t)
    
    def _build_profit_loss_table(self, story, data, style):
        """Build P&L table."""
        Paragraph = self._rl_Paragraph
        Spacer = self._rl_Spacer
        inch = self._rl_inch

        story.append(Paragraph(f'Revenue: {self._fmt(data.get("total_revenue", 0))}', style))
        story.append(Paragraph(f'COGS: {self._fmt(data.get("total_cogs", 0))}', style))
        story.append(Paragraph(f'Gross Profit: {self._fmt(data.get("gross_profit", 0))}', style))
        story.append(Spacer(1, 0.1*inch))
        story.append(Paragraph(f'Expenses: {self._fmt(data.get("total_expenses", 0))}', style))
        story.append(Paragraph(f'Net Income: {self._fmt(data.get("net_income", 0))}', style))
    
    def _build_balance_sheet_table(self, story, data, style):
        """Build Balance Sheet table."""
        Paragraph = self._rl_Paragraph

        assets = data.get('assets', {})
        story.append(Paragraph(f'Total Assets: {self._fmt(assets.get("total", 0))}', style))
        
        liabilities = data.get('liabilities', {})
        story.append(Paragraph(f'Total Liabilities: {self._fmt(liabilities.get("total", 0))}', style))
        
        equity = data.get('equity', {})
        story.append(Paragraph(f'Total Equity: {self._fmt(equity.get("total", 0))}', style))
    
    def _build_aging_table(self, story, data, style):
        """Build aging table."""
        Table = self._rl_Table
        TableStyle = self._rl_TableStyle
        colors = self._rl_colors

        table_data = [['Party', 'Current', '1-30', '31-60', '61-90', '90+', 'Total']]
        
        for row in data.get('aging_rows', []):
            table_data.append([
                row.get('party_name', '')[:20],
                self._fmt(row.get('current', 0)),
                self._fmt(row.get('age_1_30', 0)),
                self._fmt(row.get('age_31_60', 0)),
                self._fmt(row.get('age_61_90', 0)),
                self._fmt(row.get('age_90_plus', 0)),
                self._fmt(row.get('total', 0))
            ])
        
        t = Table(table_data, colWidths=[100, 60, 60, 60, 60, 60, 80])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        story.append(t)


class ExportEngine:
    """
    Unified export engine for all report types.
    Supports: CSV, Excel, PDF, JSON
    """
    
    EXPORTERS = {
        'csv': lambda: None,  # Uses ReportExporter
        'excel': ExcelExporter,
        'pdf': PDFExporter,
    }
    
    @classmethod
    def export(cls, data: dict, report_type: str, format: str = 'excel', 
                company_name: str = '') -> bytes:
        """
        Export report to specified format.
        
        Args:
            data: Report data dictionary
            report_type: Type of report (trial_balance, profit_loss, etc.)
            format: Export format (csv, excel, pdf, json)
            company_name: Company name for headers
        
        Returns:
            bytes: Export file content
        """
        format = format.lower()
        
        if format == 'csv':
            from accounting.services.report_exporter import ReportExporter
            return ReportExporter.to_csv(data, report_type).encode('utf-8')
        
        elif format == 'excel':
            return ExcelExporter().export(data, report_type, company_name)
        
        elif format == 'pdf':
            return PDFExporter().export(data, report_type, company_name)
        
        elif format == 'json':
            return json.dumps(data, indent=2, default=str).encode('utf-8')
        
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    @classmethod
    def get_available_formats(cls) -> list:
        """Get list of available export formats."""
        formats = ['csv', 'json']
        try:
            import openpyxl
            formats.append('excel')
        except ImportError:
            pass
        try:
            import reportlab
            formats.append('pdf')
        except ImportError:
            pass
        return formats