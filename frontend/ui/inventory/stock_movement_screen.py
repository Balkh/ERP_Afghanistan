"""Stock Movement screen for ERP."""
from PySide6.QtWidgets import (QVBoxLayout, QHBoxLayout,
                                  QLabel, QComboBox, QWidget,
                                  QFrame, QMessageBox, QFileDialog)
from PySide6.QtCore import Qt
from api.client import APIClient
from ui.screens.base_screen import BaseScreen, ScreenState
from ui.constants import (SPACING_XS, SPACING_SM, SPACING_MD, SPACING_LG, SPACING_XL, MARGIN_PAGE, TEXT_PAGE_TITLE,
                           TEXT_BODY, TEXT_LABEL, BORDER_RADIUS_MD, COLOR_BG_SURFACE, COLOR_BG_ELEVATED, COLOR_BORDER, COLOR_TEXT_PRIMARY, COLOR_TEXT_MUTED, COLOR_DANGER,
                           INPUT_HEIGHT_MD, TEXT_CARD_TITLE, SPACING_XS, SPACING_XXL, COLOR_SUCCESS, BORDER_RADIUS_LG)
from ui.components.buttons import EnterpriseButton, ButtonVariant, ButtonSize
from ui.components.tables import EnterpriseTable, TableColumn
from ui.components.state_helper import StateHelper
from ui.components.dialogs import EnterpriseDialog, DialogType, AlertDialog, ConfirmDialog
from api.client import APIClient
from api.endpoints import get_endpoint


class StockMovementScreen(BaseScreen):
    """Screen for managing stock movements."""

    def __init__(self, parent=None, screen_id="stock_movement", config=None, api_client=None):
        super().__init__(parent, screen_id, config)
        self.api_client = api_client or APIClient()
        self.movements = []
        self.setup_ui()
        self.load_movements()

    def setup_ui(self):
        """Setup the UI components."""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(MARGIN_PAGE, MARGIN_PAGE, MARGIN_PAGE, MARGIN_PAGE)
        layout.setSpacing(SPACING_MD + SPACING_XS)

        # Header section
        header_layout = QHBoxLayout()
        header = QLabel("Stock Movements")
        header.setStyleSheet(f"color: {COLOR_TEXT_PRIMARY}; font-size: {TEXT_PAGE_TITLE}pt; font-weight: 700;")
        header_layout.addWidget(header)

        header_layout.addStretch()

        self.btn_refresh = EnterpriseButton(text="\u27f3 Refresh", variant=ButtonVariant.SECONDARY, size=ButtonSize.MEDIUM)
        self.btn_refresh.clicked.connect(self.load_movements)
        header_layout.addWidget(self.btn_refresh)

        self.btn_export = EnterpriseButton(text="Export", variant=ButtonVariant.SECONDARY, size=ButtonSize.MEDIUM)
        self.btn_export.clicked.connect(self._export_movements)
        header_layout.addWidget(self.btn_export)

        layout.addLayout(header_layout)

        # Filters
        filter_bar = QFrame()
        filter_bar.setStyleSheet(f"""
            QFrame {{
                background-color: {COLOR_BG_ELEVATED};
                border-radius: {BORDER_RADIUS_LG};
                border: 1px solid {COLOR_BORDER};
                padding: {SPACING_MD}px;
            }}
        """)
        filter_layout = QHBoxLayout(filter_bar)
        filter_layout.setSpacing(SPACING_MD)

        filter_layout.addWidget(QLabel("Type:"))
        self.type_filter = QComboBox()
        self.type_filter.addItems(["All", "IN", "OUT", "TRANSFER", "ADJUSTMENT"])
        self.type_filter.setMinimumWidth(150)
        self.type_filter.currentTextChanged.connect(self.load_movements)
        filter_layout.addWidget(QLabel("Type:"))
        filter_layout.addWidget(self.type_filter)

        filter_layout.addWidget(QLabel("Product:"))
        self.product_search = QLineEdit()
        self.product_search.setPlaceholderText("Search product...")
        self.product_search.setMinimumHeight(INPUT_HEIGHT_MD)
        self.product_search.textChanged.connect(self.load_movements)
        filter_layout.addWidget(self.product_search)

        filter_layout.addStretch()
        layout.addWidget(filter_bar)

        # Table
        columns = [
            TableColumn("id", "ID", width=60),
            TableColumn("product_name", "Product", width=150),
            TableColumn("type", "Type", width=100, align="center"),
            TableColumn("quantity", "Qty", width=80, align="right"),
            TableColumn("warehouse_from", "From", width=120),
            TableColumn("warehouse_to", "To", width=120),
            TableColumn("reference", "Reference", width=120),
            TableColumn("date", "Date", width=100, align="center"),
            TableColumn("user", "User", width=100),
        ]
        self.table = EnterpriseTable(columns)
        layout.addWidget(self.table)

    def load_movements(self):
        """Load stock movements from API."""
        self.set_state(ScreenState.LOADING)
        try:
            endpoint = get_endpoint("stock_movements") or "/api/inventory/stock-movements/"
            params = {}
            type_filter = self.type_filter.currentText()
            if type_filter != "All":
                params["type"] = type_filter.lower()

            response = self.api_client.get(endpoint, params=params)
            movements = []
            if isinstance(response, list):
                movements = [m for m in response if isinstance(m, dict)]
            elif isinstance(response, dict) and response.get('success'):
                data = response.get('data', [])
                if isinstance(data, list):
                    movements = [m for m in data if isinstance(m, dict)]

            self.movements = movements
            self.update_table()
        except Exception as e:
            print(f"Error loading stock movements: {e}")
            self.set_state(ScreenState.ERROR)
        finally:
            self.set_state(ScreenState.READY)

    def update_table(self):
        """Update table with movement data."""
        data = []
        for move in self.movements:
            data.append({
                "id": str(move.get('id', ''))[:8],
                "product_name": move.get('product_name', ''),
                "type": move.get('type', ''),
                "quantity": f"{move.get('quantity', 0):,.0f}",
                "warehouse_from": move.get('warehouse_from_name', ''),
                "warehouse_to": move.get('warehouse_to_name', ''),
                "reference": move.get('reference', ''),
                "date": str(move.get('date', ''))[:10],
                "user": move.get('user_name', ''),
            })
        self.table.set_data(data)

    def _export_movements(self):
        """Export stock movements to Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Stock Movements"

            # Headers
            headers = ["ID", "Product", "Type", "Qty", "From Warehouse", "To Warehouse", "Reference", "Date", "User"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            # Data
            for row_idx, move in enumerate(self.movements, 2):
                ws.cell(row=row_idx, column=1, value=str(move.get('id', ''))[:8])
                ws.cell(row=row_idx, column=2, value=move.get('product_name', ''))
                ws.cell(row=row_idx, column=3, value=move.get('type', ''))
                ws.cell(row=row_idx, column=3, value=f"{move.get('quantity', 0):,.0f}")
                ws.cell(row=row_idx, column=5, value=move.get('warehouse_from_name', ''))
                ws.cell(row=row_idx, column=6, value=move.get('warehouse_to_name', ''))
                ws.cell(row=row_idx, column=7, value=move.get('reference', ''))
                ws.cell(row=row_idx, column=8, value=str(move.get('date', ''))[:10])
                ws.cell(row=row_idx, column=9, value=move.get('user_name', ''))

            # Auto-fit columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            # Save
            file_path, _ = QFileDialog.getSaveFileName(self, "Save Stock Movements Export", "stock_movements_export.xlsx", "Excel Files (*.xlsx)")
            if file_path:
                wb.save(file_path)
                AlertDialog.info(self, "Success", f"Exported to {file_path}", self)
            else:
                AlertDialog.warning(self, "Cancelled", "Export cancelled.", self)

        except ImportError:
            AlertDialog.error(self, "Error", "openpyxl library not installed. Please install with: pip install openpyxl", self)
        except Exception as e:
            AlertDialog.error(self, "Error", f"Failed to export: {e}", self)
