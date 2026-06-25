"""Payroll dialog and export utilities extracted from payroll_screen.py."""


from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                                QLineEdit, QDoubleSpinBox, QCheckBox,
                                QFileDialog)
from PySide6.QtCore import Qt
from PySide6.QtGui import QKeySequence, QShortcut
from api.endpoints import get_endpoint
from ui.constants import (SPACING_SM, INPUT_HEIGHT_LG, COLOR_PRIMARY)
from ui.components.buttons import EnterpriseButton, ButtonVariant, ButtonSize
from ui.components.dialogs import EnterpriseDialog, DialogType, AlertDialog
from ui.components.forms import FormSection
from ui.utils.validation import FormValidator


class SalaryStructureDialog(EnterpriseDialog):
    """Dialog for creating salary structures."""

    def __init__(self, parent=None, api_client=None):
        if api_client is None:
            raise ValueError("api_client is required for SalaryStructureDialog")
        super().__init__("Add Salary Structure", DialogType.CUSTOM, parent)
        self.setMinimumSize(500, 400)
        self._submitting = False
        self._structure_id = None
        self.api_client = api_client
        content = self._build_content()
        self.set_content(content)
        enter_shortcut = QShortcut(QKeySequence(Qt.Key_Return), self)
        enter_shortcut.activated.connect(self.save)
        self.name.setFocus()

    def _create_button_area(self):
        return None

    def _build_content(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)

        section = FormSection("Salary Structure", primary=True)

        self.name = QLineEdit()
        self.name.setPlaceholderText("Structure name")
        self.name.setMinimumHeight(INPUT_HEIGHT_LG)

        self.basic_salary = QDoubleSpinBox()
        self.basic_salary.setRange(0, 999999999)
        self.basic_salary.setDecimals(2)
        self.basic_salary.setValue(15000)
        self.basic_salary.setMinimumHeight(INPUT_HEIGHT_LG)

        self.is_active = QCheckBox("Active")
        self.is_active.setChecked(True)

        section.add_field(self.name, "Name*:")
        section.add_field(self.basic_salary, "Basic Salary:")
        section.add_field(self.is_active, "Status:")
        layout.addWidget(section)

        layout.addStretch()

        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(SPACING_SM)
        btn_layout.addStretch()
        cancel_btn = EnterpriseButton("Cancel", variant=ButtonVariant.SECONDARY, size=ButtonSize.MEDIUM)
        cancel_btn.clicked.connect(self.reject)
        ok_btn = EnterpriseButton("OK", variant=ButtonVariant.PRIMARY, size=ButtonSize.MEDIUM)
        ok_btn.clicked.connect(self.save)
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(ok_btn)
        layout.addLayout(btn_layout)

        return widget

    def save(self):
        if self._submitting:
            return
        self._submitting = True

        validator = FormValidator()
        validator.validate_required("Name", self.name.text().strip(), "Name is required")
        if validator.has_errors():
            error_messages = "\n".join([f"• {msg}" for msg in validator.get_errors().values()])
            AlertDialog.warning("Validation Error", f"Please fix the following errors:\n\n{error_messages}", self)
            self._submitting = False
            return

        data = {
            "name": self.name.text().strip(),
            "basic_salary": str(self.basic_salary.value()),
            "is_active": self.is_active.isChecked(),
        }
        base_endpoint = get_endpoint("payroll_records")

        def on_success(response):
            self._submitting = False
            if response and (response.get("success") or response.get("id")):
                AlertDialog.info("Success", "Salary structure saved.", self)
                self.accept()
            else:
                errors = response.get("error", "Unknown error") if isinstance(response, dict) else "Failed"
                AlertDialog.error("Error", str(errors), self)

        def on_error(message):
            self._submitting = False
            AlertDialog.error("Error", str(message), self)

        started = self.run_api_request(
            key="salary_structure_dialog_save",
            method="PUT" if self._structure_id else "POST",
            endpoint=f"{base_endpoint}{self._structure_id}/" if self._structure_id else base_endpoint,
            data=data,
            on_success=on_success,
            on_error=on_error,
        )
        if not started:
            self._submitting = False


def export_payroll_to_excel(records_table, parent=None):
    """Export payroll records to Excel using openpyxl.

    Args:
        records_table: EnterpriseTable instance with payroll data.
        parent: Parent widget for dialogs.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Payroll Records"

        headers = ["Employee", "Period", "Basic Salary", "Allowances",
                    "Deductions", "Gross", "Net", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color=COLOR_PRIMARY.lstrip("#"), end_color=COLOR_PRIMARY.lstrip("#"), fill_type="solid")

        for row_idx, record in enumerate(records_table.get_row_data(), 2):
            ws.cell(row=row_idx, column=1, value=record.get("employee_name", ""))
            ws.cell(row=row_idx, column=2, value=record.get("period", ""))
            ws.cell(row=row_idx, column=3, value=record.get("basic_salary", ""))
            ws.cell(row=row_idx, column=4, value=record.get("total_allowances", ""))
            ws.cell(row=row_idx, column=5, value=record.get("total_deductions", ""))
            ws.cell(row=row_idx, column=6, value=record.get("gross_salary", ""))
            ws.cell(row=row_idx, column=7, value=record.get("net_salary", ""))
            ws.cell(row=row_idx, column=8, value=record.get("status", ""))

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        file_path, _ = QFileDialog.getSaveFileName(
            parent, "Save Payroll Export", "payroll_export.xlsx", "Excel Files (*.xlsx)")
        if file_path:
            wb.save(file_path)
            AlertDialog.info("Success", f"Exported to {file_path}", parent)
        else:
            AlertDialog.warning("Cancelled", "Export cancelled.", parent)

    except ImportError:
        AlertDialog.error("Error",
                         "openpyxl library not installed. Please install with: pip install openpyxl",
                         parent)
    except Exception as e:
        AlertDialog.error("Error", f"Failed to export: {e}", parent)
